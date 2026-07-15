# -*- coding: utf-8 -*-
import base64
import io
import re
import unicodedata

from odoo import api, fields, models
from odoo.exceptions import UserError, ValidationError


def _slugify(text):
    value = unicodedata.normalize('NFKD', text or '').encode('ascii', 'ignore').decode('ascii')
    value = value.lower()
    value = re.sub(r'[^\w\s-]', '', value)
    value = re.sub(r'[\s_]+', '-', value)
    return value.strip('-') or 'team'


JERSEY_SIZE_SELECTION = [
    ('XS', 'XS'),
    ('S', 'S'),
    ('M', 'M'),
    ('L', 'L'),
    ('XL', 'XL'),
    ('XXL', 'XXL'),
    ('XXXL', 'XXXL'),
    ('4XL', '4XL'),
    ('5XL', '5XL'),
]

SLEEVE_SELECTION = [
    ('F', 'Full (F)'),
    ('H', 'Half (H)'),
]


class AuctionChampJerseyTeam(models.Model):
    _name = 'auction.champ.jersey.team'
    _description = 'Jersey Collection Team'
    _order = 'name'

    name = fields.Char(string='Team Name', required=True, index=True)
    slug = fields.Char(
        string='URL Slug',
        copy=False,
        index=True,
        help='Used in the public survey URL. Auto-generated from team name when empty.',
    )
    public_url = fields.Char(
        string='Public Survey URL',
        compute='_compute_public_url',
        help='Share this link to collect jersey details.',
    )
    active = fields.Boolean(default=True)
    team_logo = fields.Binary(string='Team Logo', attachment=True)
    sponsor_logo = fields.Binary(string='Sponsor Logo', attachment=True)
    jersey_design = fields.Binary(string='Jersey Design', attachment=True)
    player_ids = fields.One2many(
        'auction.champ.jersey.player',
        'team_id',
        string='Player Jersey Entries',
    )
    player_count = fields.Integer(compute='_compute_player_count', string='Entries')
    notes = fields.Text(string='Internal Notes')

    _sql_constraints = [
        ('slug_uniq', 'unique(slug)', 'The survey URL slug must be unique.'),
    ]

    @api.depends('player_ids')
    def _compute_player_count(self):
        for rec in self:
            rec.player_count = len(rec.player_ids)

    @api.depends('slug')
    def _compute_public_url(self):
        # Same pattern as player register: /<db_name>/… so multi-db prod
        # can select the correct database without a session cookie.
        base = self.env['ir.config_parameter'].sudo().get_param('web.base.url', '').rstrip('/')
        db_name = self.env.cr.dbname
        for rec in self:
            if rec.slug:
                rec.public_url = '%s/%s/auction/jersey/%s' % (base, db_name, rec.slug)
            else:
                rec.public_url = False

    @api.model
    def _unique_slug(self, base_slug, exclude_id=None):
        slug = base_slug or 'team'
        domain = [('slug', '=', slug)]
        if exclude_id:
            domain.append(('id', '!=', exclude_id))
        n = 2
        while self.sudo().search_count(domain):
            slug = '%s-%s' % (base_slug, n)
            domain = [('slug', '=', slug)]
            if exclude_id:
                domain.append(('id', '!=', exclude_id))
            n += 1
        return slug

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('slug') and vals.get('name'):
                vals['slug'] = self._unique_slug(_slugify(vals['name']))
            elif vals.get('slug'):
                vals['slug'] = self._unique_slug(_slugify(vals['slug']))
        return super().create(vals_list)

    def write(self, vals):
        vals = dict(vals)
        if vals.get('slug'):
            # Unique per record when writing shared vals to multiple records
            if len(self) == 1:
                vals['slug'] = self._unique_slug(_slugify(vals['slug']), exclude_id=self.id)
            else:
                for rec in self:
                    rec_vals = dict(vals)
                    rec_vals['slug'] = self._unique_slug(_slugify(vals['slug']), exclude_id=rec.id)
                    super(AuctionChampJerseyTeam, rec).write(rec_vals)
                return True
        return super().write(vals)

    @api.onchange('name')
    def _onchange_name_slug(self):
        if self.name and not self.slug:
            self.slug = _slugify(self.name)

    def action_open_public_url(self):
        self.ensure_one()
        if not self.public_url:
            raise ValidationError('Public URL is not available yet.')
        return {
            'type': 'ir.actions.act_url',
            'url': self.public_url,
            'target': 'new',
        }

    def action_export_excel(self):
        """Download jersey entries as Excel (data only, no logos)."""
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            raise UserError('openpyxl is not installed. Please run: pip install openpyxl')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Jersey Entries'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='0F2447')
        title_font = Font(bold=True, size=14, color='0F2447')
        thin = Border(
            left=Side(style='thin', color='D0D8E8'),
            right=Side(style='thin', color='D0D8E8'),
            top=Side(style='thin', color='D0D8E8'),
            bottom=Side(style='thin', color='D0D8E8'),
        )
        alt_fill = PatternFill('solid', fgColor='F3F6FB')
        center = Alignment(horizontal='center', vertical='center')

        ws['A1'] = 'Jersey Collection'
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        ws['A2'] = self.name or ''
        ws['A2'].font = Font(bold=True, size=12, color='1F4E79')
        ws.merge_cells('A2:E2')
        ws['A3'] = 'Total entries: %d' % len(self.player_ids)

        headers = ['#', 'Name on Jersey', 'Number', 'Size', 'Sleeve']
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        sleeve_label = {'F': 'Full (F)', 'H': 'Half (H)'}
        players = self.player_ids.sorted(key=lambda p: (p.player_name or '').lower())
        for idx, player in enumerate(players, start=1):
            row = 5 + idx
            values = [
                idx,
                player.player_name or '',
                player.number or '',
                player.size or '',
                sleeve_label.get(player.sleeve, player.sleeve or ''),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin
                if col in (1, 3, 4, 5):
                    cell.alignment = center
                if idx % 2 == 0:
                    cell.fill = alt_fill

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 10

        # Summary on the right of the same sheet (columns G–H)
        section_font = Font(bold=True, size=12, color='0F2447')
        label_font = Font(bold=True, size=10)

        ws['G5'] = 'Size Summary'
        ws['G5'].font = section_font
        ws.merge_cells('G5:H5')

        ws['G6'] = 'Size'
        ws['H6'] = 'Count'
        for col in (7, 8):
            cell = ws.cell(row=6, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        r = 7
        for sz in ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '4XL', '5XL']:
            cnt = len(players.filtered(lambda p, s=sz: p.size == s))
            if not cnt:
                continue
            c1 = ws.cell(row=r, column=7, value=sz)
            c2 = ws.cell(row=r, column=8, value=cnt)
            c1.border = thin
            c2.border = thin
            c1.alignment = center
            c2.alignment = center
            r += 1

        r += 1
        ws.cell(row=r, column=7, value='Sleeve Summary').font = section_font
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        r += 1

        for label, key in [('Full (F)', 'F'), ('Half (H)', 'H')]:
            c1 = ws.cell(row=r, column=7, value=label)
            c2 = ws.cell(row=r, column=8, value=len(players.filtered(lambda p, k=key: p.sleeve == k)))
            c1.border = thin
            c2.border = thin
            c1.font = label_font
            c2.alignment = center
            r += 1

        buf = io.BytesIO()
        wb.save(buf)
        filename = 'Jersey_%s.xlsx' % ((self.name or 'Team').replace('/', '_').replace('\\', '_').replace(' ', '_'))
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(buf.getvalue()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }


class AuctionChampJerseyPlayer(models.Model):
    _name = 'auction.champ.jersey.player'
    _description = 'Jersey Player Entry'
    _order = 'id desc'

    team_id = fields.Many2one(
        'auction.champ.jersey.team',
        string='Team',
        required=True,
        ondelete='cascade',
        index=True,
    )
    player_name = fields.Char(string='Name on Jersey', required=True)
    number = fields.Char(string='Number', help='Optional — leave blank if no jersey number is needed.')
    size = fields.Selection(JERSEY_SIZE_SELECTION, string='Size', required=True)
    sleeve = fields.Selection(SLEEVE_SELECTION, string='Sleeve', required=True)
    submitted_on = fields.Datetime(string='Submitted On', default=fields.Datetime.now, readonly=True)
