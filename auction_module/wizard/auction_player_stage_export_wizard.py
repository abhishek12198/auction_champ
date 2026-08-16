# -*- coding: utf-8 -*-
##############################################################################
#
#  AuctionChamp - Professional Sports Auction Management Platform
#
#  Copyright (c) 2026 AuctionChamp.
#  All Rights Reserved.
#
##############################################################################

import base64
import io
import re

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


# Distinct pastel fills for team colour banding in Jersey export
_TEAM_COLORS = [
    'DBEAFE',  # blue
    'DCFCE7',  # green
    'FEF3C7',  # amber
    'FCE7F3',  # pink
    'E0E7FF',  # indigo
    'FFEDD5',  # orange
    'F3E8FF',  # purple
    'CCFBF1',  # teal
    'FEE2E2',  # red
    'E2E8F0',  # slate
]
_TEAM_HEADER_COLORS = [
    '1D4ED8',
    '15803D',
    'B45309',
    'BE185D',
    '4338CA',
    'C2410C',
    '7E22CE',
    '0F766E',
    'B91C1C',
    '334155',
]

_JERSEY_SIZES = ('XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL')
_JERSEY_NAME_MAX = 20
_JERSEY_NUMBER_MAX = 4
_IMPORT_SHEET_NAME = 'Jersey Export'
_GUIDELINES_SHEET_NAME = 'Guidelines'
_PLAYER_ID_HEADER = 'Player ID'


class AuctionPlayerStageExportWizard(models.TransientModel):
    _name = 'auction.player.stage.export.wizard'
    _description = 'Export Players to Excel'

    tournament_id = fields.Many2one(
        'auction.tournament', string='Tournament', required=True, readonly=True)
    tournament_type = fields.Selection(
        related='tournament_id.tournament_type', string='Sport', readonly=True)
    export_scope = fields.Selection(
        [
            ('draft', 'Registered (Draft)'),
            ('sold', 'Sold'),
            ('unsold', 'Unsold'),
            ('sold_unsold', 'Sold and Unsold'),
            ('jersey', 'Jersey Export'),
        ],
        string='Export',
        default='sold_unsold',
        required=True,
    )
    draft_count = fields.Integer(string='Registered Players', compute='_compute_counts')
    sold_count = fields.Integer(string='Sold Players', compute='_compute_counts')
    unsold_count = fields.Integer(string='Unsold Players', compute='_compute_counts')
    jersey_count = fields.Integer(string='Jersey Players', compute='_compute_counts')
    file_data = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='Filename', readonly=True)
    import_file = fields.Binary(string='Jersey Excel to Import')
    import_file_name = fields.Char(string='Import Filename')
    import_log = fields.Text(string='Import Result', readonly=True)
    state = fields.Selection(
        [
            ('choose', 'Choose'),
            ('done', 'Done'),
            ('import_done', 'Import Done'),
        ],
        default='choose',
    )

    @api.depends('tournament_id')
    def _compute_counts(self):
        Player = self.env['auction.team.player']
        for wiz in self:
            tid = wiz.tournament_id.id
            if not tid:
                wiz.draft_count = 0
                wiz.sold_count = 0
                wiz.unsold_count = 0
                wiz.jersey_count = 0
                continue
            base = [('tournament_id', '=', tid), ('icon_player', '=', False)]
            wiz.draft_count = Player.search_count(base + [('state', '=', 'draft')])
            wiz.sold_count = Player.search_count(base + [('state', '=', 'sold')])
            wiz.unsold_count = Player.search_count(base + [('state', '=', 'unsold')])
            wiz.jersey_count = Player.search_count(
                base + [('state', 'in', ('draft', 'auction', 'sold', 'unsold'))]
            )

    def action_generate_excel(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_(
                'openpyxl is not installed. Please run: pip install openpyxl'
            ))
        if not self.tournament_id:
            raise UserError(_('Please select a tournament.'))

        scope = self.export_scope
        if scope == 'draft' and not self.draft_count:
            raise UserError(_('No registered (draft) players found for this tournament.'))
        if scope == 'sold' and not self.sold_count:
            raise UserError(_('No sold players found for this tournament.'))
        if scope == 'unsold' and not self.unsold_count:
            raise UserError(_('No unsold players found for this tournament.'))
        if scope == 'sold_unsold' and not self.sold_count and not self.unsold_count:
            raise UserError(_('No sold or unsold players found for this tournament.'))
        if scope == 'jersey' and not self.jersey_count:
            raise UserError(_('No players found for jersey export.'))

        wb = Workbook()
        default = wb.active
        wb.remove(default)

        if scope == 'jersey':
            self._write_jersey_workbook(wb)
        elif scope == 'draft':
            self._write_sheet(wb, 'Registered', 'draft')
        elif scope == 'sold':
            self._write_sheet(wb, 'Sold', 'sold')
        elif scope == 'unsold':
            self._write_sheet(wb, 'Unsold', 'unsold')
        else:
            if self.sold_count:
                self._write_sheet(wb, 'Sold', 'sold')
            if self.unsold_count:
                self._write_sheet(wb, 'Unsold', 'unsold')

        if not wb.sheetnames:
            raise UserError(_('Nothing to export for the selected option.'))

        buf = io.BytesIO()
        wb.save(buf)
        safe = re.sub(r'[^\w\-]+', '_', self.tournament_id.name or 'tournament')
        fname = 'players_%s_%s.xlsx' % (safe, scope)
        self.write({
            'file_data': base64.b64encode(buf.getvalue()),
            'file_name': fname,
            'import_log': False,
            'state': 'done',
        })
        return self._reopen()

    def action_import_jersey(self):
        """Update only jersey name / number / size from the protected import sheet."""
        self.ensure_one()
        if load_workbook is None:
            raise UserError(_(
                'openpyxl is not installed. Please run: pip install openpyxl'
            ))
        if not self.tournament_id:
            raise UserError(_('Please select a tournament.'))
        if not self.import_file:
            raise UserError(_(
                'Please upload the Jersey Export Excel file after editing '
                'only the Jersey Name / Number / Size columns.'
            ))

        try:
            wb = load_workbook(
                filename=io.BytesIO(base64.b64decode(self.import_file)),
                data_only=True,
            )
        except Exception as exc:
            raise UserError(_('Could not read Excel file: %s') % exc) from exc

        ws = self._find_jersey_import_sheet(wb)
        header_row_idx, header_map = self._locate_jersey_header_row(ws)
        if _PLAYER_ID_HEADER.lower() not in header_map:
            raise UserError(_(
                'The uploaded file must contain a "%s" column '
                '(use the Jersey Export Excel from this wizard).'
            ) % _PLAYER_ID_HEADER)

        pid_col = header_map[_PLAYER_ID_HEADER.lower()]
        name_col = header_map.get('jersey name')
        number_col = header_map.get('jersey number')
        size_col = header_map.get('jersey size')
        if name_col is None and number_col is None and size_col is None:
            raise UserError(_(
                'Could not find Jersey Name / Jersey Number / Jersey Size columns.'
            ))

        Player = self.env['auction.team.player'].sudo()
        updated = 0
        unchanged = 0
        skipped = 0
        errors = []

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            pid_raw = ws.cell(row=row_idx, column=pid_col + 1).value
            if pid_raw is None or str(pid_raw).strip() == '':
                # Skip banner / spacer / empty rows
                continue
            try:
                player_id = int(float(str(pid_raw).strip()))
            except (TypeError, ValueError):
                skipped += 1
                continue

            player = Player.browse(player_id).exists()
            if not player or player.tournament_id.id != self.tournament_id.id:
                errors.append(
                    'Row %d: Player ID %s not found in this tournament.'
                    % (row_idx, player_id)
                )
                continue
            if player.icon_player:
                skipped += 1
                continue

            try:
                j_name = self._normalize_jersey_name(
                    self._cell_str(ws.cell(row=row_idx, column=name_col + 1).value)
                    if name_col is not None else (player.jersy_name or '')
                )
                j_number = self._normalize_jersey_number(
                    self._cell_str(ws.cell(row=row_idx, column=number_col + 1).value)
                    if number_col is not None else (player.jersy_number or '')
                )
                j_size = self._normalize_jersey_size(
                    self._cell_str(ws.cell(row=row_idx, column=size_col + 1).value)
                    if size_col is not None else (player.jersy_size or '')
                )
            except UserError as exc:
                errors.append('Row %d (%s): %s' % (
                    row_idx, player.name or player_id, exc.args[0] if exc.args else exc))
                continue

            vals = {}
            if (player.jersy_name or '') != j_name:
                vals['jersy_name'] = j_name or False
            if (player.jersy_number or '') != j_number:
                vals['jersy_number'] = j_number or False
            if (player.jersy_size or '') != j_size:
                vals['jersy_size'] = j_size or False

            if vals:
                player.write(vals)
                updated += 1
            else:
                unchanged += 1

        log_lines = [
            'Jersey import finished for: %s' % (self.tournament_id.name or ''),
            'Updated: %d' % updated,
            'Unchanged: %d' % unchanged,
            'Skipped rows: %d' % skipped,
        ]
        if errors:
            log_lines.append('Issues (%d):' % len(errors))
            log_lines.extend(errors[:50])
            if len(errors) > 50:
                log_lines.append('… and %d more.' % (len(errors) - 50))

        self.write({
            'import_log': '\n'.join(log_lines),
            'state': 'import_done',
            'file_data': False,
            'file_name': False,
        })
        return self._reopen()

    def action_back(self):
        self.ensure_one()
        self.write({
            'state': 'choose',
            'file_data': False,
            'file_name': False,
            'import_file': False,
            'import_file_name': False,
            'import_log': False,
        })
        return self._reopen()

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ── Styles ────────────────────────────────────────────────────────────

    def _header_style(self, fg='1B3F8F'):
        return (
            PatternFill('solid', fgColor=fg),
            Font(color='FFFFFF', bold=True, size=11),
            Alignment(horizontal='center', vertical='center', wrap_text=True),
            Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            ),
        )

    def _thin_border(self):
        return Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

    def _editable_fill(self):
        return PatternFill('solid', fgColor='FEF9C3')  # soft yellow

    def _editable_header_fill(self):
        return PatternFill('solid', fgColor='CA8A04')

    # ── Column definitions ────────────────────────────────────────────────

    def _common_headers(self):
        return ['Sl.No', 'Player Name', 'Mobile Number', 'Org ID#', 'Card No']

    def _sport_headers(self):
        t = self.tournament_id
        if t.tournament_type == 'football':
            cols = [
                'Playing Position',
                'Secondary Positions',
                'Preferred Foot',
                'Age',
                'Height',
                'Weight',
                'Work Rate',
                'Playing Styles',
                'Strengths',
            ]
            for lab in t.other_attribute_label_ids.sorted('sequence'):
                if lab.label and lab.label not in cols:
                    cols.append(lab.label)
            return cols
        return [
            'Role',
            'Batting Style',
            'Bowling Style',
            'Blood Group',
        ]

    def _jersey_headers(self, include_player_id=False):
        headers = []
        if include_player_id:
            headers.append(_PLAYER_ID_HEADER)
        return headers + self._common_headers() + self._sport_headers() + [
            'Jersey Name', 'Jersey Number', 'Jersey Size', 'Team', 'Status',
        ]

    def _jersey_editable_headers(self):
        return {'Jersey Name', 'Jersey Number', 'Jersey Size'}

    def _stage_headers(self, stage):
        base = self._common_headers() + self._sport_headers()
        if stage == 'sold':
            return base + ['Sold To', 'Sold Points', 'Base Point', 'Tier',
                           'Jersey Name', 'Jersey Number', 'Jersey Size']
        if stage == 'draft':
            return base + ['Status', 'Tier', 'Base Point',
                           'Jersey Name', 'Jersey Number', 'Jersey Size']
        return base + ['Status', 'Tier', 'Base Point',
                       'Jersey Name', 'Jersey Number', 'Jersey Size']

    def _sold_points_map(self, players):
        if not players:
            return {}
        lines = self.env['auction.auction.player'].sudo().search(
            [('player_id', 'in', players.ids)], order='id asc')
        mapping = {}
        for line in lines:
            mapping[line.player_id.id] = line.points or 0
        return mapping

    def _sport_values(self, player):
        t = self.tournament_id
        if t.tournament_type == 'football':
            vals = [
                player.dominant_position_id.name if player.dominant_position_id else '',
                ', '.join(player.secondary_position_ids.mapped('name')),
                (player.preferred_foot or '').capitalize(),
                player.age or '',
                player.height or '',
                player.weight or '',
                (player.work_rate or '').capitalize(),
                ', '.join(player.playing_style_ids.mapped('name')),
                ', '.join(player.strength_ids.mapped('name')),
            ]
            attr_map = {
                (a.label or '').strip().lower(): (a.value or '').strip()
                for a in player.other_attribute_ids
            }
            for lab in t.other_attribute_label_ids.sorted('sequence'):
                if not lab.label:
                    continue
                vals.append(attr_map.get(lab.label.strip().lower(), ''))
            return vals
        return [
            player.role or '',
            player.batting_style or '',
            player.bowling_style or '',
            player.blood_group or '',
        ]

    def _jersey_values(self, player):
        return [
            (player.jersy_name or '').upper(),
            player.jersy_number or '',
            player.jersy_size or '',
        ]

    def _state_label(self, state):
        return {
            'draft': 'Registered',
            'auction': 'In Auction',
            'sold': 'Sold',
            'unsold': 'Unsold',
        }.get(state, state or '')

    def _player_base_row(self, seq, player, include_player_id=False):
        row = []
        if include_player_id:
            row.append(player.id)
        row.extend([
            seq,
            (player.name or '').upper(),
            player.contact or '',
            player.org_id or '',
            player.sl_no or '',
        ])
        row.extend(self._sport_values(player))
        return row

    # ── Jersey field normalisation (import) ───────────────────────────────

    @staticmethod
    def _cell_str(val):
        if val is None:
            return ''
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val).strip()

    def _normalize_jersey_name(self, value):
        name = (value or '').strip().upper()
        # Normalise curly quotes / dashes from Excel
        name = (name
                .replace('\u2019', "'").replace('\u2018', "'")
                .replace('\u201c', '"').replace('\u201d', '"')
                .replace('\u2013', '-').replace('\u2014', '-'))
        name = name.replace('"', '')
        if not name:
            return ''
        if len(name) > _JERSEY_NAME_MAX:
            raise UserError(_(
                'Jersey Name must be at most %d characters (got %d).'
            ) % (_JERSEY_NAME_MAX, len(name)))
        # Letters, spaces, hyphen, apostrophe, period only
        if not re.fullmatch(r"[A-Z0-9 .'\-]+", name):
            raise UserError(_(
                'Jersey Name may only contain letters, numbers, spaces, '
                "hyphen (-), apostrophe ('), and period (.)."
            ))
        return name

    def _normalize_jersey_number(self, value):
        number = (value or '').strip().upper()
        if not number:
            return ''
        if len(number) > _JERSEY_NUMBER_MAX:
            raise UserError(_(
                'Jersey Number must be at most %d characters.'
            ) % _JERSEY_NUMBER_MAX)
        if not re.fullmatch(r'[A-Z0-9]+', number):
            raise UserError(_(
                'Jersey Number may only contain letters and digits (no spaces).'
            ))
        return number

    def _normalize_jersey_size(self, value):
        size = (value or '').strip().upper().replace(' ', '')
        if not size:
            return ''
        # Accept "XL — Extra Large" style from old exports
        size = size.split('—')[0].split('-')[0].strip()
        aliases = {
            '2XL': 'XXL',
            '3XL': 'XXXL',
            'EXTRA SMALL': 'XS',
            'SMALL': 'S',
            'MEDIUM': 'M',
            'LARGE': 'L',
            'EXTRA LARGE': 'XL',
            'DOUBLE XL': 'XXL',
            'TRIPLE XL': 'XXXL',
        }
        size = aliases.get(size, size)
        if size not in _JERSEY_SIZES:
            raise UserError(_(
                'Jersey Size must be one of: %s'
            ) % ', '.join(_JERSEY_SIZES))
        return size

    def _find_jersey_import_sheet(self, wb):
        if _IMPORT_SHEET_NAME in wb.sheetnames:
            return wb[_IMPORT_SHEET_NAME]
        # Fallback: first sheet that has Player ID header
        for name in wb.sheetnames:
            if name == _GUIDELINES_SHEET_NAME:
                continue
            ws = wb[name]
            for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row or 1), values_only=True):
                headers = [str(h).strip().lower() if h is not None else '' for h in row]
                if _PLAYER_ID_HEADER.lower() in headers:
                    return ws
        raise UserError(_(
            'Could not find the "%s" sheet. Please upload the Jersey Export '
            'Excel generated from this wizard.'
        ) % _IMPORT_SHEET_NAME)

    def _locate_jersey_header_row(self, ws):
        for row_idx in range(1, min(20, (ws.max_row or 1) + 1)):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
            header_map = {}
            for idx, h in enumerate(row):
                if h is None:
                    continue
                key = str(h).strip().lower()
                if key and key not in header_map:
                    header_map[key] = idx
            if _PLAYER_ID_HEADER.lower() in header_map:
                return row_idx, header_map
        raise UserError(_('Could not locate the header row with "%s".') % _PLAYER_ID_HEADER)

    # ── Stage sheets (Registered / Sold / Unsold) ─────────────────────────

    def _write_sheet(self, wb, title, stage):
        ws = wb.create_sheet(title)
        headers = self._stage_headers(stage)
        header_colors = {
            'sold': '15803D',
            'unsold': 'B91C1C',
            'draft': '1D4ED8',
        }
        fill, font, align, border = self._header_style(header_colors.get(stage, '1B3F8F'))

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                12, min(28, len(header) + 4))

        players = self.env['auction.team.player'].search(
            [
                ('tournament_id', '=', self.tournament_id.id),
                ('state', '=', stage),
                ('icon_player', '=', False),
            ],
            order='sl_no asc, name asc',
        )
        points_map = self._sold_points_map(players) if stage == 'sold' else {}
        thin = self._thin_border()

        for seq, player in enumerate(players, start=1):
            row = self._player_base_row(seq, player)
            if stage == 'sold':
                row.extend([
                    player.assigned_team_id.name if player.assigned_team_id else '',
                    points_map.get(player.id, 0),
                    player.effective_base_price or 0,
                    player.tier_id.name if player.tier_id else '',
                ])
            else:
                row.extend([
                    self._state_label(stage),
                    player.tier_id.name if player.tier_id else '',
                    player.effective_base_price or 0,
                ])
            row.extend(self._jersey_values(player))

            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=seq + 1, column=col_idx, value=value)
                cell.border = thin
                if col_idx == 2:
                    cell.font = Font(bold=True)

        ws.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(headers)), max(1, len(players) + 1))
        ws.freeze_panes = 'A2'

    # ── Jersey export: 1 main sheet + Guidelines (sheet 2) ───────────────

    def _write_jersey_workbook(self, wb):
        players = self.env['auction.team.player'].search(
            [
                ('tournament_id', '=', self.tournament_id.id),
                ('icon_player', '=', False),
                ('state', 'in', ('draft', 'auction', 'sold', 'unsold')),
            ],
            order='assigned_team_id, sl_no asc, name asc',
        )
        # Sheet 1 = data, Sheet 2 = guidelines
        self._write_jersey_main_sheet(wb, players)
        self._write_jersey_guidelines_sheet(wb)

    def _write_jersey_guidelines_sheet(self, wb):
        ws = wb.create_sheet(_GUIDELINES_SHEET_NAME)  # append as sheet 2
        ws.sheet_view.showGridLines = False

        title_fill = PatternFill('solid', fgColor='0F172A')
        section_fill = PatternFill('solid', fgColor='1E3A5F')
        tip_fill = PatternFill('solid', fgColor='FEF3C7')
        ok_fill = PatternFill('solid', fgColor='DCFCE7')
        bad_fill = PatternFill('solid', fgColor='FEE2E2')

        ws.merge_cells('A1:B1')
        c = ws['A1']
        c.value = 'Jersey Export — Guidelines'
        c.fill = title_fill
        c.font = Font(color='FFFFFF', bold=True, size=16)
        c.alignment = Alignment(vertical='center', horizontal='left')
        ws.row_dimensions[1].height = 32

        ws.merge_cells('A2:B2')
        ws['A2'].value = (
            'Tournament: %s  ·  Edit ONLY the yellow columns on "%s", then Import Jersey Updates.'
            % (self.tournament_id.name or '', _IMPORT_SHEET_NAME)
        )
        ws['A2'].font = Font(color='475569', italic=True, size=10)

        rows = [
            (4, 'HOW TO USE', section_fill, True),
            (5, '1. On "%s", edit only the yellow columns: Jersey Name, Jersey Number, Jersey Size.'
               % _IMPORT_SHEET_NAME, None, False),
            (6, '2. Do not change Player ID (needed to match players on import).', None, False),
            (7, '3. Other columns (name, mobile, attributes, team, status) are ignored on import.', None, False),
            (8, '4. Save the file → Export Players → Import Jersey Updates.', None, False),
            (10, 'JERSEY NAME', section_fill, True),
            (11, '• Name printed on the jersey (usually short name / surname).', None, False),
            (12, '• Use UPPERCASE. Maximum %d characters.' % _JERSEY_NAME_MAX, None, False),
            (13, "• Allowed: A–Z, 0–9, space, hyphen (-), apostrophe ('), period (.).", None, False),
            (14, "Examples OK: RAHUL  ·  D'SOUZA  ·  AL-AMIN", ok_fill, False),
            (15, 'Avoid: emoji, @ # $ %, full sentences.', bad_fill, False),
            (17, 'JERSEY NUMBER', section_fill, True),
            (18, '• Squad / jersey number on the kit.', None, False),
            (19, '• Maximum %d characters. Letters and digits only (no spaces).' % _JERSEY_NUMBER_MAX, None, False),
            (20, 'Examples OK: 7  ·  10  ·  18  ·  99', ok_fill, False),
            (21, 'Avoid: 1000 (too long), values with spaces, #18.', bad_fill, False),
            (23, 'JERSEY SIZE — type exactly one of these codes', section_fill, True),
        ]

        for row_idx, text, fill, is_section in rows:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = ws.cell(row=row_idx, column=1, value=text)
            if is_section:
                cell.fill = fill
                cell.font = Font(color='FFFFFF', bold=True, size=11)
                ws.row_dimensions[row_idx].height = 20
            else:
                if fill:
                    cell.fill = fill
                cell.font = Font(size=10, color='0F172A')

        # Size table (code | meaning)
        size_rows = [
            ('XS', 'Extra Small'),
            ('S', 'Small'),
            ('M', 'Medium'),
            ('L', 'Large'),
            ('XL', 'Extra Large'),
            ('XXL', 'Double XL'),
            ('XXXL', 'Triple XL'),
        ]
        ws['A24'] = 'Code'
        ws['B24'] = 'Meaning'
        ws['A24'].fill = tip_fill
        ws['B24'].fill = tip_fill
        ws['A24'].font = Font(bold=True, size=10)
        ws['B24'].font = Font(bold=True, size=10)

        for i, (code, meaning) in enumerate(size_rows):
            r = 25 + i
            ws.cell(row=r, column=1, value=code).font = Font(bold=True, size=11)
            ws.cell(row=r, column=1).fill = ok_fill
            ws.cell(row=r, column=2, value=meaning).font = Font(size=10)

        note_row = 25 + len(size_rows) + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
        ws.cell(
            row=note_row, column=1,
            value='Type the Code exactly (e.g. XL). Leave blank only if size is unknown.',
        ).fill = tip_fill

        ws.column_dimensions['A'].width = 72
        ws.column_dimensions['B'].width = 22

    def _write_jersey_main_sheet(self, wb, players):
        """Single data sheet: team banners + coloured rows; yellow jersey cols editable."""
        ws = wb.create_sheet(_IMPORT_SHEET_NAME, 0)
        headers = self._jersey_headers(include_player_id=True)
        editable = self._jersey_editable_headers()
        thin = self._thin_border()
        edit_fill = self._editable_fill()
        edit_header_fill = self._editable_header_fill()
        ncol = len(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        banner = ws.cell(
            row=1, column=1,
            value=(
                'EDIT YELLOW COLUMNS ONLY — Jersey Name / Number / Size  ·  '
                'Players grouped by Assigned Team (colour bands)  ·  '
                'Size codes → Guidelines sheet  ·  Other columns ignored on import'
            ),
        )
        banner.fill = PatternFill('solid', fgColor='854D0E')
        banner.font = Font(color='FFFFFF', bold=True, size=11)
        banner.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 26

        # Column headers (row 2)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            if header in editable:
                cell.fill = edit_header_fill
                cell.font = Font(color='FFFFFF', bold=True, size=11)
            else:
                fill, font, align, border = self._header_style('1E3A5F')
                cell.fill = fill
                cell.font = font
                cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin
            width = 12 if header == _PLAYER_ID_HEADER else max(12, min(28, len(header) + 4))
            if header in editable:
                width = max(width, 16)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        pid_col = headers.index(_PLAYER_ID_HEADER) + 1

        # Build team groups: assigned teams (A–Z), then unassigned
        teams = players.mapped('assigned_team_id').sorted(lambda t: (t.name or '').lower())
        groups = []
        for idx, team in enumerate(teams):
            team_players = players.filtered(
                lambda p, tid=team.id: p.assigned_team_id.id == tid
            ).sorted(lambda p: (p.sl_no or 0, p.name or ''))
            if team_players:
                groups.append((team, team_players, idx % len(_TEAM_COLORS)))
        unassigned = players.filtered(lambda p: not p.assigned_team_id).sorted(
            lambda p: (p.sl_no or 0, p.name or ''))
        if unassigned:
            groups.append((None, unassigned, None))

        row_idx = 3
        global_seq = 0

        for team, team_players, color_idx in groups:
            if team:
                header_fg = _TEAM_HEADER_COLORS[color_idx]
                row_fill = PatternFill('solid', fgColor=_TEAM_COLORS[color_idx])
                banner_text = '▶  TEAM: %s    (%d players)' % (
                    (team.name or 'Team').upper(), len(team_players))
            else:
                header_fg = '475569'
                row_fill = PatternFill('solid', fgColor='E2E8F0')
                banner_text = '▶  UNASSIGNED / NO TEAM    (%d players)' % len(team_players)

            # Team section banner (no Player ID → skipped on import)
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx, end_column=ncol)
            team_banner = ws.cell(row=row_idx, column=1, value=banner_text)
            team_banner.fill = PatternFill('solid', fgColor=header_fg)
            team_banner.font = Font(color='FFFFFF', bold=True, size=12)
            team_banner.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1

            for player in team_players:
                global_seq += 1
                values = (
                    self._player_base_row(global_seq, player, include_player_id=True)
                    + self._jersey_values(player)
                    + [
                        player.assigned_team_id.name if player.assigned_team_id else '',
                        self._state_label(player.state),
                    ]
                )
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin
                    header = headers[col_idx - 1]
                    if header in editable:
                        cell.fill = edit_fill
                        cell.alignment = Alignment(
                            horizontal='left' if header == 'Jersey Name' else 'center')
                    elif col_idx == pid_col:
                        cell.fill = PatternFill('solid', fgColor='F8FAFC')
                        cell.font = Font(color='64748B', size=9)
                    else:
                        cell.fill = row_fill
                        if header == 'Player Name':
                            cell.font = Font(bold=True)
                        # Left accent bar feel on Sl.No
                        if header == 'Sl.No':
                            cell.fill = PatternFill('solid', fgColor=header_fg)
                            cell.font = Font(color='FFFFFF', bold=True)
                            cell.alignment = Alignment(horizontal='center')
                row_idx += 1

            # Spacer between teams
            row_idx += 1

        ws.freeze_panes = 'C3'
