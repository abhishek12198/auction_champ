# -*- coding: utf-8 -*-
##############################################################################
#
#  AuctionChamp - Professional Sports Auction Management Platform
#
#  Copyright (c) 2026 AuctionChamp.
#  All Rights Reserved.
#
#  CONFIDENTIAL & PROPRIETARY
#
#  This source code, including but not limited to its algorithms, business
#  logic, database structures, models, controllers, views, reports, templates,
#  APIs, documentation, and related materials, constitutes proprietary and
#  confidential information owned exclusively by AuctionChamp.
#
#  This software is protected by applicable copyright laws and international
#  intellectual property treaties. Unauthorized copying, reproduction,
#  modification, distribution, publication, sublicensing, reverse engineering,
#  decompilation, disassembly, disclosure, or use of this software, in whole
#  or in part, is strictly prohibited without the prior written permission of
#  AuctionChamp.
#
#  This software is licensed, not sold. Possession of the source code does not
#  grant any right to copy, modify, redistribute, or create derivative works
#  except as expressly permitted under a valid written license agreement with
#  AuctionChamp.
#
#  Any unauthorized use may result in civil and criminal penalties under
#  applicable intellectual property and copyright laws.
#
#  Company  : AuctionChamp
#  Website  : www.auctionchamp.live
#  Email    : auctionchamp.live@gmail.com
#
#  © 2026 AuctionChamp. All Rights Reserved.
#
##############################################################################

import base64
import io
import os
import re
import zipfile

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


_PHOTO_EXTS = {'.jpg', '.jpeg', '.png'}


class AuctionPlayerUploadWizard(models.TransientModel):
    _name = 'auction.player.upload.wizard'
    _description = 'Upload Players from Excel + Photo ZIP'

    tournament_id = fields.Many2one(
        'auction.tournament', string='Tournament', required=True, readonly=True)
    tournament_type = fields.Selection(
        related='tournament_id.tournament_type', readonly=True)
    excel_file = fields.Binary(string='Players Excel (.xlsx)')
    excel_filename = fields.Char(string='Excel Filename')
    photos_zip = fields.Binary(
        string='Player Photos ZIP',
        help='Optional. ZIP of photos named 1.jpg, 2.png, … matching Excel row order '
             '(first data row = 1). Supports .jpg / .jpeg / .png (any case).')
    photos_zip_filename = fields.Char(string='ZIP Filename')
    template_file = fields.Binary(string='Download Template', readonly=True)
    template_filename = fields.Char(readonly=True)
    state = fields.Selection(
        [('draft', 'Upload'), ('done', 'Done')], default='draft')
    result_message = fields.Text(string='Result', readonly=True)
    created_count = fields.Integer(readonly=True)
    photo_count = fields.Integer(readonly=True)

    def _cricket_columns(self):
        return [
            'Serial No',
            'Name',
            'Contact',
            'Role',
            'Batting Style',
            'Bowling Style',
            'Blood Group',
            'Category',
            'Location',
            'Tier',
            'Base Price',
            'Jersey Name',
            'Jersey Number',
            'Jersey Size',
        ]

    def _football_base_columns(self):
        return [
            'Serial No',
            'Name',
            'Contact',
            'Playing Position',
            'Secondary Positions',
            'Preferred Foot',
            'Age',
            'Height',
            'Weight',
            'Playing Styles',
            'Strengths',
            'Work Rate',
            'Blood Group',
            'Category',
            'Location',
            'Tier',
            'Base Price',
            'Jersey Name',
            'Jersey Number',
            'Jersey Size',
        ]

    def _excel_columns(self):
        self.ensure_one()
        if self.tournament_id.tournament_type == 'football':
            cols = list(self._football_base_columns())
            for lab in self.tournament_id.other_attribute_label_ids.sorted('sequence'):
                if lab.label and lab.label not in cols:
                    cols.append(lab.label)
            return cols
        return list(self._cricket_columns())

    def action_download_template(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_(
                'openpyxl is not installed. Please run: pip install openpyxl'
            ))
        columns = self._excel_columns()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Players'

        header_fill = PatternFill('solid', fgColor='1B3F8F')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )
        for col_idx, title in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(title) + 4))

        sample = {c: '' for c in columns}
        sample['Serial No'] = '1'
        sample['Name'] = 'Sample Player'
        sample['Contact'] = '9876543210'
        if self.tournament_id.tournament_type == 'football':
            sample['Playing Position'] = 'CB'
            sample['Secondary Positions'] = 'CB, RB'
            sample['Preferred Foot'] = 'Left'
            sample['Age'] = '25'
            sample['Playing Styles'] = 'Target Man'
            sample['Strengths'] = 'Speed'
            sample['Work Rate'] = 'Medium'
            for lab in self.tournament_id.other_attribute_label_ids:
                if lab.label in sample:
                    sample[lab.label] = 'Example'
        else:
            sample['Role'] = 'Batsman'
            sample['Batting Style'] = 'Right Handed Batter'
            sample['Bowling Style'] = 'Right Arm'
        sample['Blood Group'] = 'A+'
        sample['Tier'] = 'Rookie'
        for col_idx, title in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=sample.get(title, ''))
            cell.border = thin
            cell.font = Font(italic=True, color='888888')

        self._write_field_guide_sheet(wb)
        self._write_reference_lists_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        sport = self.tournament_id.tournament_type or 'players'
        fname = 'player_upload_%s_%s.xlsx' % (
            re.sub(r'[^\w\-]+', '_', self.tournament_id.name or 'tournament'),
            sport,
        )
        self.write({
            'template_file': base64.b64encode(buf.getvalue()),
            'template_filename': fname,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _write_field_guide_sheet(self, wb):
        """Column-by-column data entry guide (type + how to fill)."""
        ws = wb.create_sheet('Field Guide', 1)
        header_fill = PatternFill('solid', fgColor='1B3F8F')
        header_font = Font(color='FFFFFF', bold=True)
        tip_fill = PatternFill('solid', fgColor='FFF8E1')
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        sport = (self.tournament_id.tournament_type or '').title()
        intro = [
            'Player Upload — Field Guide | %s | %s' % (self.tournament_id.name, sport),
            'Use this sheet before filling Players. See "Reference Lists" for exact values from your database.',
            'Legend — Free text: type anything. Selection: only listed options. Many2one: one matching name. '
            'Many2many: one or more names, comma-separated.',
            'Name is required. All other columns are optional. Row 2 on Players is a sample — replace or delete it.',
            'Photos ZIP (optional): files named 1.jpg, 2.png, 3.jpeg … matching Excel row order '
            '(first data row after header = 1). PNG / JPG / JPEG supported.',
        ]
        for i, line in enumerate(intro, start=1):
            cell = ws.cell(row=i, column=1, value=line)
            cell.fill = tip_fill
            cell.font = Font(bold=(i == 1))
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

        headers = ['Column', 'Field Type', 'How to Enter', 'Allowed Values / Notes']
        header_row = len(intro) + 2
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        rows = self._field_guide_rows()
        for r_idx, row in enumerate(rows, start=header_row + 1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if c_idx == 2:
                    cell.font = Font(bold=True, color='1B3F8F')

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 72
        ws.row_dimensions[header_row].height = 22

    def _field_guide_rows(self):
        """Return list of [column, type, how, notes] for current sport."""
        common_start = [
            [
                'Serial No',
                'Free text / Number',
                'Optional. Leave blank to auto-number in Excel order.',
                'Used for display order and to match photo filenames (1.jpg = first data row).',
            ],
            [
                'Name',
                'Free text',
                'Required. Player full name.',
                'Do not leave blank. Rows starting with "Sample" are skipped on import.',
            ],
            [
                'Contact',
                'Free text / Number',
                'Phone / mobile number as text or number.',
                'Optional. Stored as text on the player.',
            ],
        ]
        common_end = [
            [
                'Blood Group',
                'Free text',
                'Type as text (e.g. A+, B-, O+, AB+).',
                'Not a dropdown in Excel — enter the usual blood-group string.',
            ],
            [
                'Category',
                'Free text',
                'Any category label used by your tournament.',
                'Free text (e.g. Local, Outstation). Not linked to a master list.',
            ],
            [
                'Location',
                'Free text',
                'City / place name.',
                'Optional free text.',
            ],
            [
                'Tier',
                'Many2one',
                'Enter the exact Tier Name for THIS tournament (one value only).',
                'Must match a tier on this tournament (see Reference Lists → Tiers). '
                'Matching is case-insensitive. Wrong name = import error for that row.',
            ],
            [
                'Base Price',
                'Number',
                'Numeric base / starting price points.',
                'Example: 1000. Leave blank if not used.',
            ],
            [
                'Jersey Name',
                'Free text',
                'Name printed on jersey.',
                'Optional.',
            ],
            [
                'Jersey Number',
                'Free text / Number',
                'Squad / jersey number.',
                'Optional.',
            ],
            [
                'Jersey Size',
                'Free text',
                'Size label (e.g. S, M, L, XL).',
                'Free text — not a fixed selection in upload.',
            ],
        ]

        if self.tournament_id.tournament_type == 'football':
            mid = [
                [
                    'Playing Position',
                    'Many2one',
                    'ONE primary position. Use Position Name or short Code.',
                    'Must exist in Player Positions master (see Reference Lists). '
                    'Examples: Centre Back or CB. Case-insensitive.',
                ],
                [
                    'Secondary Positions',
                    'Many2many',
                    'Zero or more positions, comma-separated. Name or Code for each.',
                    'Example: CB, RB, LB  or  Centre Back, Right Back. '
                    'See Reference Lists → Positions. Unknown values are skipped with a warning.',
                ],
                [
                    'Preferred Foot',
                    'Selection',
                    'Exactly one of the fixed options (case-insensitive).',
                    'Allowed only: Left | Right | Both',
                ],
                [
                    'Age',
                    'Number',
                    'Whole number age in years.',
                    'Example: 25',
                ],
                [
                    'Height',
                    'Number',
                    'Height value (as used in your process, e.g. cm).',
                    'Numeric. Example: 178',
                ],
                [
                    'Weight',
                    'Number',
                    'Weight value (e.g. kg).',
                    'Numeric. Example: 72',
                ],
                [
                    'Playing Styles',
                    'Many2many',
                    'Zero or more styles, comma-separated. Use exact Style Name.',
                    'Must match Playing Styles master (see Reference Lists). Example: Target Man, Playmaker',
                ],
                [
                    'Strengths',
                    'Many2many',
                    'Zero or more strengths, comma-separated. Use exact Strength Name.',
                    'Must match Strengths master (see Reference Lists). Example: Speed, Stamina',
                ],
                [
                    'Work Rate',
                    'Selection',
                    'Exactly one of the fixed options (case-insensitive).',
                    'Allowed only: Low | Medium | High',
                ],
            ]
            for lab in self.tournament_id.other_attribute_label_ids.sorted('sequence'):
                if not lab.label:
                    continue
                mid.append([
                    lab.label,
                    'Free text (Other Attribute)',
                    'Type the Label-Value for this Att-Label on the player.',
                    'Column comes from tournament Other Attributes → Att-Labels. '
                    'Whatever you type becomes the value for "%s". Leave blank to skip.' % lab.label,
                ])
            return common_start + mid + common_end

        mid = [
            [
                'Role',
                'Free text',
                'Playing role as text (not a fixed dropdown in upload).',
                'Examples often used: Batsman, Bowler, All Rounder, Wicket Keeper. '
                'Any string is accepted.',
            ],
            [
                'Batting Style',
                'Free text',
                'Describe batting style in plain text.',
                'Examples: Right Handed Batter, Left Handed Batter. Free text.',
            ],
            [
                'Bowling Style',
                'Free text',
                'Describe bowling style in plain text.',
                'Examples: Right Arm Fast, Left Arm Orthodox, Leg Spin. Free text.',
            ],
        ]
        return common_start + mid + common_end

    def _write_reference_lists_sheet(self, wb):
        """Live allowed values from master / this tournament."""
        ws = wb.create_sheet('Reference Lists', 2)
        header_fill = PatternFill('solid', fgColor='0D7377')
        header_font = Font(color='FFFFFF', bold=True)
        section_fill = PatternFill('solid', fgColor='E0F2F1')
        tip_fill = PatternFill('solid', fgColor='FFF8E1')
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        def section_title(row, title):
            cell = ws.cell(row=row, column=1, value=title)
            cell.fill = section_fill
            cell.font = Font(bold=True, size=12)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            return row + 1

        def write_table(row, headers, data_rows):
            for c_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=c_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin
            row += 1
            if not data_rows:
                cell = ws.cell(row=row, column=1, value='(none found)')
                cell.font = Font(italic=True, color='888888')
                return row + 2
            for data in data_rows:
                for c_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=c_idx, value=val)
                    cell.border = thin
                row += 1
            return row + 1

        row = 1
        note = ws.cell(
            row=row, column=1,
            value=(
                'Copy values from these lists into the Players sheet. '
                'Names must match (case-insensitive). Do not invent codes/names that are not listed.'
            ),
        )
        note.fill = tip_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        row = 3

        # Tiers (tournament-specific Many2one)
        row = section_title(row, 'TIERS (Many2one) — use Tier Name in the Tier column')
        tiers = self.env['auction.player.tier'].search([
            ('tournament_id', '=', self.tournament_id.id),
        ])
        row = write_table(
            row,
            ['Tier Name (enter this)', 'Description', 'Icon Tier'],
            [
                [
                    t.name or '',
                    t.description or '',
                    'Yes' if t.is_an_icon_tier else '',
                ]
                for t in tiers
            ],
        )

        if self.tournament_id.tournament_type == 'football':
            row = section_title(
                row,
                'PLAYING POSITIONS (Many2one / Many2many) — use Name OR Code',
            )
            positions = self.env['auction.player.position'].search([])
            row = write_table(
                row,
                ['Position Name', 'Code (short)', 'Enter either'],
                [
                    [
                        p.name or '',
                        p.code or '',
                        '%s  or  %s' % (p.name or '', p.code or p.name or ''),
                    ]
                    for p in positions
                ],
            )

            row = section_title(
                row,
                'PREFERRED FOOT (Selection) — only these values',
            )
            row = write_table(
                row,
                ['Allowed Value', 'Meaning'],
                [
                    ['Left', 'Left foot preferred'],
                    ['Right', 'Right foot preferred'],
                    ['Both', 'Both feet'],
                ],
            )

            row = section_title(
                row,
                'WORK RATE (Selection) — only these values',
            )
            row = write_table(
                row,
                ['Allowed Value'],
                [['Low'], ['Medium'], ['High']],
            )

            row = section_title(
                row,
                'PLAYING STYLES (Many2many) — comma-separated Style Names',
            )
            styles = self.env['auction.player.style'].search([])
            row = write_table(
                row,
                ['Style Name (enter this)'],
                [[s.name or ''] for s in styles],
            )

            row = section_title(
                row,
                'STRENGTHS (Many2many) — comma-separated Strength Names',
            )
            strengths = self.env['auction.player.strength'].search([])
            row = write_table(
                row,
                ['Strength Name (enter this)'],
                [[s.name or ''] for s in strengths],
            )

            row = section_title(
                row,
                'OTHER ATTRIBUTE COLUMNS (Free text Label-Values) — this tournament',
            )
            labels = self.tournament_id.other_attribute_label_ids.sorted('sequence')
            row = write_table(
                row,
                ['Column Header (Att-Label)', 'What to type'],
                [
                    [
                        lab.label or '',
                        'Any text value for this label (becomes Other Attribute on the player)',
                    ]
                    for lab in labels
                ] if labels else [],
            )
        else:
            row = section_title(row, 'CRICKET — Role / Batting Style / Bowling Style')
            row = write_table(
                row,
                ['Column', 'Type', 'Suggested examples (free text — not enforced)'],
                [
                    ['Role', 'Free text', 'Batsman, Bowler, All Rounder, Wicket Keeper'],
                    ['Batting Style', 'Free text', 'Right Handed Batter, Left Handed Batter'],
                    ['Bowling Style', 'Free text', 'Right Arm Fast, Left Arm Orthodox, Leg Spin, Off Spin'],
                ],
            )

        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 48

    def action_import(self):
        self.ensure_one()
        if load_workbook is None:
            raise UserError(_(
                'openpyxl is not installed. Please run: pip install openpyxl'
            ))
        if not self.excel_file:
            raise UserError(_('Please upload a filled Players Excel (.xlsx) file.'))

        try:
            wb = load_workbook(
                filename=io.BytesIO(base64.b64decode(self.excel_file)),
                data_only=True,
            )
        except Exception as exc:
            raise UserError(_('Could not read Excel file: %s') % exc) from exc

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('The Excel file is empty.'))

        headers = [self._norm_header(h) for h in rows[0]]
        if not any(headers):
            raise UserError(_('The Excel header row is missing.'))

        header_map = {}
        for idx, h in enumerate(headers):
            if h and h not in header_map:
                header_map[h] = idx

        name_key = self._find_header(header_map, ['name', 'player name', 'player'])
        if name_key is None:
            raise UserError(_('Excel must include a "Name" column.'))

        photo_map = self._parse_photos_zip()
        Player = self.env['auction.team.player'].sudo()
        created = 0
        photos_applied = 0
        errors = []
        warnings = []
        skipped_over_limit = []
        data_row_no = 0
        tid = self.tournament_id.id
        max_players, plan_name = self._saas_max_players_per_tournament()
        current_count = Player.search_count([
            ('tournament_id', '=', tid),
            ('icon_player', '=', False),
        ])

        for raw in rows[1:]:
            if not raw or all(v is None or str(v).strip() == '' for v in raw):
                continue
            data_row_no += 1

            def cell(key_aliases, _raw=raw, _map=header_map):
                key = self._find_header(_map, key_aliases)
                if key is None:
                    return ''
                val = _raw[_map[key]] if _map[key] < len(_raw) else None
                if val is None:
                    return ''
                if isinstance(val, float) and val.is_integer():
                    return str(int(val))
                return str(val).strip()

            name = cell(['name', 'player name', 'player'])
            if not name:
                data_row_no -= 1
                continue
            if name.lower().startswith('sample'):
                continue

            try:
                if max_players is not None and current_count >= max_players:
                    skipped_over_limit.append(name)
                    continue
                vals = self._row_to_player_vals(cell, name, data_row_no, header_map, raw)
                photo_b64 = photo_map.get(data_row_no)
                if photo_b64:
                    vals['photo'] = photo_b64
                Player.with_context(
                    default_tournament_id=tid,
                    saas_skip_quota_check=True,
                ).create(vals)
                created += 1
                current_count += 1
                if photo_b64:
                    photos_applied += 1
            except Exception as exc:
                errors.append('Row %s (%s): %s' % (data_row_no, name, exc))

        msg_lines = [
            'Imported %s player(s) into %s.' % (created, self.tournament_id.name),
            'Photos attached: %s.' % photos_applied,
        ]
        if skipped_over_limit:
            warnings.append(
                'Your %(plan)s plan allows up to %(max)s player(s) per tournament. '
                '"%(tourn)s" reached that limit — %(n)s player(s) were not created: %(names)s.'
                % {
                    'plan': plan_name or 'current',
                    'max': max_players,
                    'tourn': self.tournament_id.name,
                    'n': len(skipped_over_limit),
                    'names': ', '.join(skipped_over_limit[:12])
                             + ('…' if len(skipped_over_limit) > 12 else ''),
                }
            )
        if warnings:
            msg_lines.append('')
            msg_lines.append('Warnings:')
            msg_lines.extend(warnings)
        if errors:
            msg_lines.append('')
            msg_lines.append('Skipped / failed rows:')
            msg_lines.extend(errors[:30])
            if len(errors) > 30:
                msg_lines.append('… and %s more.' % (len(errors) - 30))

        self.write({
            'state': 'done',
            'created_count': created,
            'photo_count': photos_applied,
            'result_message': '\n'.join(msg_lines),
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _saas_max_players_per_tournament(self):
        """Plan player cap for this tournament, or (None, None) if unlimited."""
        if 'ac.saas.account' not in self.env:
            return None, None
        Account = self.env['ac.saas.account']
        account = Account._get_account_for_user()
        if not account and 'saas_account_id' in self.tournament_id._fields:
            account = self.tournament_id.saas_account_id
        if not account or not account.plan_id:
            return None, None
        plan = account.plan_id
        return plan.max_players_per_tournament, plan.name

    def action_close(self):
        return {'type': 'ir.actions.act_window_close'}

    @staticmethod
    def _norm_header(value):
        if value is None:
            return ''
        return re.sub(r'\s+', ' ', str(value).strip()).lower()

    @staticmethod
    def _find_header(header_map, aliases):
        for alias in aliases:
            key = alias.lower().strip()
            if key in header_map:
                return key
        return None

    def _parse_photos_zip(self):
        """Return {1: base64, 2: base64, …} from ZIP entries named N.ext."""
        self.ensure_one()
        mapping = {}
        if not self.photos_zip:
            return mapping
        try:
            zf = zipfile.ZipFile(io.BytesIO(base64.b64decode(self.photos_zip)))
        except Exception as exc:
            raise UserError(_('Could not read photos ZIP: %s') % exc) from exc

        for info in zf.infolist():
            if info.is_dir():
                continue
            basename = os.path.basename(info.filename)
            if not basename or basename.startswith('.') or '__MACOSX' in info.filename:
                continue
            stem, ext = os.path.splitext(basename)
            if ext.lower() not in _PHOTO_EXTS:
                continue
            match = re.search(r'(\d+)$', stem.strip())
            if not match:
                continue
            num = int(match.group(1))
            if num < 1:
                continue
            mapping[num] = base64.b64encode(zf.read(info)).decode('ascii')
        return mapping

    def _resolve_tier(self, tier_name):
        if not tier_name:
            return False
        Tier = self.env['auction.player.tier']
        domain_extra = []
        if 'tournament_id' in Tier._fields:
            domain_extra = ['|', ('tournament_id', '=', self.tournament_id.id),
                            ('tournament_id', '=', False)]
        tier = Tier.search([('name', '=ilike', tier_name)] + domain_extra, limit=1)
        if not tier:
            tier = Tier.search([('name', '=ilike', tier_name)], limit=1)
        return tier.id if tier else False

    def _resolve_positions(self, text, multi=False):
        Position = self.env['auction.player.position']
        if not text:
            return [] if multi else False
        parts = [p.strip() for p in re.split(r'[,;/|]+', text) if p.strip()]
        ids = []
        for part in parts:
            pos = Position.search([
                '|', ('code', '=ilike', part), ('name', '=ilike', part)
            ], limit=1)
            if pos:
                ids.append(pos.id)
        if multi:
            return ids
        return ids[0] if ids else False

    def _resolve_m2m_by_name(self, model, text):
        if not text:
            return []
        Model = self.env[model]
        ids = []
        for part in [p.strip() for p in re.split(r'[,;/|]+', text) if p.strip()]:
            rec = Model.search([('name', '=ilike', part)], limit=1)
            if rec:
                ids.append(rec.id)
        return ids

    def _row_to_player_vals(self, cell, name, data_row_no, header_map, raw):
        tournament = self.tournament_id
        sl_raw = cell(['serial no', 'serial', 'sl no', 'sl_no', '#'])
        try:
            sl_no = int(float(sl_raw)) if sl_raw else data_row_no
        except ValueError:
            sl_no = data_row_no

        vals = {
            'tournament_id': tournament.id,
            'name': name,
            'sl_no': sl_no,
            'contact': cell(['contact', 'mobile', 'phone', 'mobile no', 'mobile number']),
            'blood_group': cell(['blood group', 'blood']),
            'p_category': cell(['category', 'player category']),
            'address': cell(['location', 'address', 'venue']),
            'jersy_name': cell(['jersey name', 'jersy name', 'name on jersey']),
            'jersy_number': cell(['jersey number', 'jersy number', 'jersey no']),
            'jersy_size': cell(['jersey size', 'jersy size']),
            'state': 'draft',
        }
        tier_id = self._resolve_tier(cell(['tier']))
        if tier_id:
            vals['tier_id'] = tier_id
        base_price = cell(['base price', 'base_price', 'base point'])
        if base_price:
            try:
                vals['base_price'] = int(float(base_price))
            except ValueError:
                pass

        if tournament.tournament_type == 'football':
            pos_id = self._resolve_positions(
                cell(['playing position', 'position', 'dominant position']), multi=False)
            if pos_id:
                vals['dominant_position_id'] = pos_id
            sec_ids = self._resolve_positions(
                cell(['secondary positions', 'secondary', 'secondary position']), multi=True)
            if sec_ids:
                vals['secondary_position_ids'] = [(6, 0, sec_ids)]
            foot = cell(['preferred foot', 'foot']).lower()
            if foot in ('left', 'right', 'both'):
                vals['preferred_foot'] = foot
            age = cell(['age'])
            if age:
                try:
                    vals['age'] = int(float(age))
                except ValueError:
                    pass
            height = cell(['height'])
            weight = cell(['weight'])
            if height:
                vals['height'] = height
            if weight:
                vals['weight'] = weight
            style_ids = self._resolve_m2m_by_name(
                'auction.player.style',
                cell(['playing styles', 'playing style', 'styles']))
            if style_ids:
                vals['playing_style_ids'] = [(6, 0, style_ids)]
            strength_ids = self._resolve_m2m_by_name(
                'auction.player.strength',
                cell(['strengths', 'strength']))
            if strength_ids:
                vals['strength_ids'] = [(6, 0, strength_ids)]
            rate = cell(['work rate', 'workrate']).lower()
            if rate in ('low', 'medium', 'high'):
                vals['work_rate'] = rate

            attr_commands = []
            for lab in tournament.other_attribute_label_ids.sorted('sequence'):
                label = (lab.label or '').strip()
                if not label:
                    continue
                key = self._norm_header(label)
                value = ''
                if key in header_map:
                    idx = header_map[key]
                    raw_val = raw[idx] if idx < len(raw) else None
                    if raw_val is not None:
                        if isinstance(raw_val, float) and raw_val.is_integer():
                            value = str(int(raw_val))
                        else:
                            value = str(raw_val).strip()
                attr_commands.append((0, 0, {
                    'label': label,
                    'value': value,
                    'sequence': lab.sequence or 10,
                }))
            if attr_commands:
                vals['other_attribute_ids'] = attr_commands
        else:
            role = cell(['role'])
            if role:
                vals['role'] = role
            batting = cell(['batting style', 'batting'])
            bowling = cell(['bowling style', 'bowling'])
            if batting:
                vals['batting_style'] = batting
            if bowling:
                vals['bowling_style'] = bowling

        return vals
