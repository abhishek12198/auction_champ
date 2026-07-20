# -*- coding: utf-8 -*-
##############################################################################
#
#  AuctionChamp - Professional Sports Auction Management Platform
#
#  Copyright (c) 2026 AuctionChamp.
#  All Rights Reserved.
#
##############################################################################

import base64
import io
import os
import re
import zipfile

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


_LOGO_EXTS = {'.jpg', '.jpeg', '.png'}

# Color selection on auction.player.tier: hex → label
_TIER_COLORS = [
    ('#e74c3c', 'Red'),
    ('#e67e22', 'Orange'),
    ('#f39c12', 'Yellow'),
    ('#2ecc71', 'Green'),
    ('#1abc9c', 'Teal'),
    ('#3498db', 'Blue'),
    ('#2980b9', 'Dark Blue'),
    ('#9b59b6', 'Purple'),
    ('#e91e63', 'Pink'),
    ('#34495e', 'Dark'),
    ('#7f8c8d', 'Gray'),
    ('#ffffff', 'White'),
]


class AuctionTeamUploadWizard(models.TransientModel):
    _name = 'auction.team.upload.wizard'
    _description = 'Upload Teams & Tiers from Excel + Logo ZIP'

    tournament_id = fields.Many2one(
        'auction.tournament', string='Tournament', required=True, readonly=True)
    excel_file = fields.Binary(string='Excel (.xlsx)')
    excel_filename = fields.Char(string='Excel Filename')
    logos_zip = fields.Binary(
        string='Team Logos ZIP',
        help='Optional. ZIP of logos named 1.jpg, 2.png, … matching Excel Sl.No / row order. '
             'Supports .jpg / .jpeg / .png (any case).')
    logos_zip_filename = fields.Char(string='ZIP Filename')
    template_file = fields.Binary(string='Download Template', readonly=True)
    template_filename = fields.Char(readonly=True)
    state = fields.Selection(
        [('draft', 'Upload'), ('done', 'Done')], default='draft')
    result_message = fields.Text(string='Result', readonly=True)
    created_count = fields.Integer(string='Teams Created', readonly=True)
    updated_count = fields.Integer(string='Teams Updated', readonly=True)
    logo_count = fields.Integer(readonly=True)
    tier_created_count = fields.Integer(string='Tiers Created', readonly=True)
    tier_updated_count = fields.Integer(string='Tiers Updated', readonly=True)

    def _team_columns(self):
        return ['Sl.No', 'Team Name', 'Owner Name']

    def _tier_columns(self):
        return [
            'Tier Name',
            'Description',
            'Color',
            'Icon Tier',
            'Mystery',
        ]

    def action_download_template(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_(
                'openpyxl is not installed. Please run: pip install openpyxl'
            ))
        wb = Workbook()
        self._write_teams_sheet(wb)
        self._write_tiers_sheet(wb)
        self._write_instructions_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        safe = re.sub(r'[^\w\-]+', '_', self.tournament_id.name or 'tournament')
        fname = 'team_tier_upload_%s.xlsx' % safe
        self.write({
            'template_file': base64.b64encode(buf.getvalue()),
            'template_filename': fname,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _header_style(self):
        return (
            PatternFill('solid', fgColor='1B3F8F'),
            Font(color='FFFFFF', bold=True),
            Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            ),
        )

    def _write_teams_sheet(self, wb):
        ws = wb.active
        ws.title = 'Teams'
        header_fill, header_font, thin = self._header_style()
        columns = self._team_columns()
        for col_idx, title in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(title) + 4))

        samples = [
            {'Sl.No': 1, 'Team Name': 'Sample Warriors', 'Owner Name': 'Ravi Kumar'},
            {'Sl.No': 2, 'Team Name': 'Sample Strikers', 'Owner Name': 'Anil Mehta'},
        ]
        for row_idx, sample in enumerate(samples, start=2):
            for col_idx, title in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sample.get(title, ''))
                cell.border = thin
                cell.font = Font(italic=True, color='888888')

    def _write_tiers_sheet(self, wb):
        ws = wb.create_sheet('Tiers')
        header_fill, header_font, thin = self._header_style()
        columns = self._tier_columns()
        for col_idx, title in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = max(14, min(22, len(title) + 6))

        samples = [
            {
                'Tier Name': 'Sample Rookie',
                'Description': 'Entry level',
                'Color': 'Blue',
                'Icon Tier': 'No',
                'Mystery': 'No',
            },
            {
                'Tier Name': 'Sample Icon',
                'Description': 'Star players',
                'Color': 'Red',
                'Icon Tier': 'Yes',
                'Mystery': 'No',
            },
            {
                'Tier Name': 'Sample Mystery',
                'Description': 'Hidden until sold',
                'Color': 'Purple',
                'Icon Tier': 'No',
                'Mystery': 'Yes',
            },
        ]
        for row_idx, sample in enumerate(samples, start=2):
            for col_idx, title in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sample.get(title, ''))
                cell.border = thin
                cell.font = Font(italic=True, color='888888')

    def _write_instructions_sheet(self, wb):
        ws = wb.create_sheet('Instructions')
        tip_fill = PatternFill('solid', fgColor='FFF8E1')
        header_fill = PatternFill('solid', fgColor='0D7377')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        tips = [
            'Teams & Tiers Upload — %s' % (self.tournament_id.name or 'Tournament'),
            '',
            'SHEETS',
            '• Teams — create/update teams (logos via optional ZIP).',
            '• Tiers — create/update player tiers for this tournament (color + checkboxes).',
            '• You can fill either sheet or both. Sample rows (name starts with Sample) are skipped.',
            '',
            'TEAMS',
            '1. Team Name is required. Owner Name is optional.',
            '2. Sl.No links each row to a logo in the ZIP (1.jpg, 2.png, …).',
            '3. If Sl.No is blank, logos match Excel data-row order (first data row = 1).',
            '4. Same Team Name in this tournament → update Owner / Logo (no duplicate).',
            '',
            'TIERS — column guide',
        ]
        for i, line in enumerate(tips, start=1):
            cell = ws.cell(row=i, column=1, value=line)
            if i == 1:
                cell.font = Font(bold=True)
                cell.fill = tip_fill

        guide = [
            ['Column', 'Type', 'How to Enter', 'Notes'],
            [
                'Tier Name',
                'Free text',
                'Required. Unique name for this tournament.',
                'Same name → update that tier (case-insensitive).',
            ],
            [
                'Description',
                'Free text',
                'Optional short description.',
                '',
            ],
            [
                'Color',
                'Selection',
                'Color label OR hex code (see table below).',
                'Default Blue if blank/invalid.',
            ],
            [
                'Icon Tier',
                'Checkbox (Yes/No)',
                'Yes / No / True / False / 1 / 0',
                'Only ONE Icon Tier per tournament. Last Yes in the sheet wins.',
            ],
            [
                'Mystery',
                'Checkbox (Yes/No)',
                'Yes / No / True / False / 1 / 0',
                'Mystery players hide details on live stage until sold.',
            ],
        ]
        start = len(tips) + 2
        for r_idx, row in enumerate(guide):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=start + r_idx, column=c_idx, value=val)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if r_idx == 0:
                    cell.fill = header_fill
                    cell.font = header_font

        color_start = start + len(guide) + 2
        ws.cell(row=color_start, column=1, value='ALLOWED COLORS (enter Label or Hex)').font = Font(bold=True)
        ws.cell(row=color_start + 1, column=1, value='Label').fill = header_fill
        ws.cell(row=color_start + 1, column=1).font = header_font
        ws.cell(row=color_start + 1, column=2, value='Hex').fill = header_fill
        ws.cell(row=color_start + 1, column=2).font = header_font
        for i, (hex_code, label) in enumerate(_TIER_COLORS):
            ws.cell(row=color_start + 2 + i, column=1, value=label).border = thin
            ws.cell(row=color_start + 2 + i, column=2, value=hex_code).border = thin

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 48

    def action_import(self):
        self.ensure_one()
        if load_workbook is None:
            raise UserError(_(
                'openpyxl is not installed. Please run: pip install openpyxl'
            ))
        if not self.excel_file:
            raise UserError(_('Please upload a filled Excel (.xlsx) file.'))

        try:
            wb = load_workbook(
                filename=io.BytesIO(base64.b64decode(self.excel_file)),
                data_only=True,
            )
        except Exception as exc:
            raise UserError(_('Could not read Excel file: %s') % exc) from exc

        teams_ws = self._find_sheet(wb, ['teams', 'team'])
        tiers_ws = self._find_sheet(wb, ['tiers', 'tier', 'player tiers'])
        # Backward compatible: old templates had only a Teams-style active sheet
        if teams_ws is None and tiers_ws is None:
            teams_ws = wb.active

        logo_map = self._parse_logos_zip()
        team_created = team_updated = logos_applied = 0
        tier_created = tier_updated = 0
        errors = []
        warnings = []

        if teams_ws is not None:
            tc, tu, la, terr, twarn = self._import_teams_sheet(teams_ws, logo_map)
            team_created, team_updated, logos_applied = tc, tu, la
            errors.extend(terr)
            warnings.extend(twarn)

        if tiers_ws is not None:
            tc, tu, terr = self._import_tiers_sheet(tiers_ws)
            tier_created, tier_updated = tc, tu
            errors.extend(terr)

        msg_lines = [
            'Imported into %s.' % self.tournament_id.name,
            'Teams — Created: %s. Updated: %s. Logos attached: %s.' % (
                team_created, team_updated, logos_applied),
            'Tiers — Created: %s. Updated: %s.' % (tier_created, tier_updated),
        ]
        if warnings:
            msg_lines.append('')
            msg_lines.append('Warnings:')
            msg_lines.extend(warnings[:20])
        if errors:
            msg_lines.append('')
            msg_lines.append('Skipped / failed rows:')
            msg_lines.extend(errors[:40])
            if len(errors) > 40:
                msg_lines.append('… and %s more.' % (len(errors) - 40))

        self.write({
            'state': 'done',
            'created_count': team_created,
            'updated_count': team_updated,
            'logo_count': logos_applied,
            'tier_created_count': tier_created,
            'tier_updated_count': tier_updated,
            'result_message': '\n'.join(msg_lines),
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    @staticmethod
    def _find_sheet(wb, aliases):
        wanted = {a.lower().strip() for a in aliases}
        for name in wb.sheetnames:
            if name.lower().strip() in wanted:
                return wb[name]
        return None

    def _saas_max_teams_per_tournament(self):
        """Plan team cap for this tournament, or None if unlimited / no SaaS."""
        if 'ac.saas.account' not in self.env:
            return None, None
        Account = self.env['ac.saas.account']
        account = Account._get_account_for_user()
        if not account:
            # Tournament may still be owned by a SaaS account
            owner = self.tournament_id.saas_account_id if (
                'saas_account_id' in self.tournament_id._fields
            ) else False
            account = owner
        if not account or not account.plan_id:
            return None, None
        plan = account.plan_id
        return plan.max_teams_per_tournament, plan.name

    def _import_teams_sheet(self, ws, logo_map):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0, 0, 0, [], []

        headers = [self._norm_header(h) for h in rows[0]]
        if not any(headers):
            return 0, 0, 0, [], []

        header_map = {}
        for idx, h in enumerate(headers):
            if h and h not in header_map:
                header_map[h] = idx

        if self._find_header(header_map, ['team name', 'name', 'team']) is None:
            return 0, 0, 0, [
                'Teams sheet: missing "Team Name" column — teams not imported.',
            ], []

        Team = self.env['auction.team'].sudo()
        created = updated = logos_applied = 0
        errors = []
        warnings = []
        skipped_over_limit = []
        data_row_no = 0
        tid = self.tournament_id.id
        max_teams, plan_name = self._saas_max_teams_per_tournament()
        current_count = Team.search_count([('tournament_id', '=', tid)])

        for raw in rows[1:]:
            if not raw or all(v is None or str(v).strip() == '' for v in raw):
                continue
            data_row_no += 1

            def cell(key_aliases, _raw=raw, _map=header_map):
                return self._cell_value(_raw, _map, key_aliases)

            name = cell(['team name', 'name', 'team'])
            if not name:
                data_row_no -= 1
                continue
            if name.lower().startswith('sample'):
                continue

            owner = cell(['owner name', 'owner', 'manager', 'team owner'])
            sl_raw = cell(['sl.no', 'sl no', 'serial no', 'serial', 'sl_no', '#'])
            try:
                sl_no = int(float(sl_raw)) if sl_raw else data_row_no
            except ValueError:
                sl_no = data_row_no

            try:
                vals = {
                    'tournament_id': tid,
                    'name': name,
                    'manager': owner or False,
                }
                logo_b64 = logo_map.get(sl_no)
                if logo_b64:
                    vals['logo'] = logo_b64

                existing = Team.search([
                    ('tournament_id', '=', tid),
                    ('name', '=ilike', name),
                ], limit=1)
                if existing:
                    existing.write(vals)
                    updated += 1
                    if logo_b64:
                        logos_applied += 1
                else:
                    if max_teams is not None and current_count >= max_teams:
                        skipped_over_limit.append(name)
                        continue
                    Team.with_context(
                        default_tournament_id=tid,
                        saas_skip_quota_check=True,
                    ).create(vals)
                    created += 1
                    current_count += 1
                    if logo_b64:
                        logos_applied += 1
            except Exception as exc:
                errors.append('Teams row %s (%s): %s' % (data_row_no, name, exc))

        if skipped_over_limit:
            warnings.append(
                'Your %(plan)s plan allows up to %(max)s team(s) per tournament. '
                '"%(tourn)s" already has / reached that limit — '
                '%(n)s new team(s) were not created: %(names)s. '
                'Existing teams in the sheet were still updated.'
                % {
                    'plan': plan_name or 'current',
                    'max': max_teams,
                    'tourn': self.tournament_id.name,
                    'n': len(skipped_over_limit),
                    'names': ', '.join(skipped_over_limit[:12])
                             + ('…' if len(skipped_over_limit) > 12 else ''),
                }
            )

        return created, updated, logos_applied, errors, warnings

    def _import_tiers_sheet(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0, 0, []

        headers = [self._norm_header(h) for h in rows[0]]
        if not any(headers):
            return 0, 0, []

        header_map = {}
        for idx, h in enumerate(headers):
            if h and h not in header_map:
                header_map[h] = idx

        if self._find_header(header_map, [
            'tier name', 'name', 'tier',
        ]) is None:
            return 0, 0, [
                'Tiers sheet: missing "Tier Name" column — tiers not imported.',
            ]

        Tier = self.env['auction.player.tier']
        created = updated = 0
        errors = []
        data_row_no = 0
        # Collect icon-tier candidates; last Yes wins
        icon_tier_name = None
        pending = []  # list of (name, vals without icon flag resolution)

        for raw in rows[1:]:
            if not raw or all(v is None or str(v).strip() == '' for v in raw):
                continue
            data_row_no += 1

            def cell(key_aliases, _raw=raw, _map=header_map):
                return self._cell_value(_raw, _map, key_aliases)

            name = cell(['tier name', 'name', 'tier'])
            if not name:
                data_row_no -= 1
                continue
            if name.lower().startswith('sample'):
                continue

            description = cell(['description', 'desc'])
            color_raw = cell(['color', 'tier color'])
            color = self._resolve_tier_color(color_raw)
            icon_flag = self._parse_bool(cell([
                'icon tier', 'is an icon tier', 'icon', 'is_an_icon_tier',
            ]))
            mystery_flag = self._parse_bool(cell([
                'mystery', 'mystery tier', 'is mystery',
            ]))

            if color_raw and color is None:
                errors.append(
                    'Tiers row %s (%s): unknown Color "%s" — using Blue. '
                    'Allowed: %s' % (
                        data_row_no, name, color_raw,
                        ', '.join(label for _, label in _TIER_COLORS),
                    )
                )
                color = '#3498db'

            vals = {
                'tournament_id': self.tournament_id.id,
                'name': name,
                'description': description or False,
                'color': color or '#3498db',
                'mystery': mystery_flag,
                # Set icon after all rows so only one is True
                'is_an_icon_tier': False,
            }
            if icon_flag:
                icon_tier_name = name
            pending.append((data_row_no, name, vals))

        for data_row_no, name, vals in pending:
            try:
                existing = Tier.search([
                    ('tournament_id', '=', self.tournament_id.id),
                    ('name', '=ilike', name),
                ], limit=1)
                if existing:
                    existing.write(vals)
                    updated += 1
                else:
                    Tier.create(vals)
                    created += 1
            except Exception as exc:
                errors.append('Tiers row %s (%s): %s' % (data_row_no, name, exc))

        # Apply single Icon Tier for this tournament
        if icon_tier_name:
            try:
                others = Tier.search([
                    ('tournament_id', '=', self.tournament_id.id),
                    ('is_an_icon_tier', '=', True),
                ])
                if others:
                    others.write({'is_an_icon_tier': False})
                icon_rec = Tier.search([
                    ('tournament_id', '=', self.tournament_id.id),
                    ('name', '=ilike', icon_tier_name),
                ], limit=1)
                if icon_rec:
                    icon_rec.write({'is_an_icon_tier': True})
            except Exception as exc:
                errors.append('Icon Tier "%s": %s' % (icon_tier_name, exc))

        return created, updated, errors

    def _resolve_tier_color(self, raw):
        """Map label or hex to selection key. Return None if unknown (and raw set)."""
        if not raw:
            return '#3498db'
        text = str(raw).strip()
        lower = text.lower()
        # Hex with or without #
        hex_candidate = lower if lower.startswith('#') else ('#%s' % lower)
        for hex_code, label in _TIER_COLORS:
            if lower == label.lower() or hex_candidate == hex_code.lower():
                return hex_code
        return None

    @staticmethod
    def _parse_bool(raw):
        if raw is None or raw == '':
            return False
        if isinstance(raw, bool):
            return raw
        text = str(raw).strip().lower()
        if text in ('1', 'true', 'yes', 'y', 'on', 'checked'):
            return True
        if text in ('0', 'false', 'no', 'n', 'off', ''):
            return False
        # Excel sometimes stores True as 1.0 already stringified upstream
        return False

    @staticmethod
    def _cell_value(raw, header_map, key_aliases):
        key = AuctionTeamUploadWizard._find_header(header_map, key_aliases)
        if key is None:
            return ''
        val = raw[header_map[key]] if header_map[key] < len(raw) else None
        if val is None:
            return ''
        if isinstance(val, bool):
            return 'Yes' if val else 'No'
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val).strip()

    @staticmethod
    def _norm_header(value):
        if value is None:
            return ''
        return re.sub(r'\s+', ' ', str(value).strip()).lower()

    @staticmethod
    def _find_header(header_map, aliases):
        for alias in aliases:
            key = alias.lower().strip()
            if key in header_map:
                return key
        return None

    def _parse_logos_zip(self):
        """Return {1: base64, 2: base64, …} from ZIP entries named N.ext."""
        self.ensure_one()
        mapping = {}
        if not self.logos_zip:
            return mapping
        try:
            zf = zipfile.ZipFile(io.BytesIO(base64.b64decode(self.logos_zip)))
        except Exception as exc:
            raise UserError(_('Could not read logos ZIP: %s') % exc) from exc

        for info in zf.infolist():
            if info.is_dir():
                continue
            basename = os.path.basename(info.filename)
            if not basename or basename.startswith('.') or '__MACOSX' in info.filename:
                continue
            stem, ext = os.path.splitext(basename)
            if ext.lower() not in _LOGO_EXTS:
                continue
            match = re.search(r'(\d+)$', stem.strip())
            if not match:
                continue
            num = int(match.group(1))
            if num < 1:
                continue
            mapping[num] = base64.b64encode(zf.read(info)).decode('ascii')
        return mapping
